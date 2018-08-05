#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Excel使用

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(sheet['A1'].value)
c = sheet['B1']
s = sheet.cell(row=1, column=2)
print(s)

for i in range(1, 8, 2):
    print(i, sheet.cell(row = i, column = 2).value)

c = get_column_letter(1)
print(c)
b = get_column_letter(27)
print(b)
a = column_index_from_string('A')
print(a)